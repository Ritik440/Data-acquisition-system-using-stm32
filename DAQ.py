import os
from tkinter import filedialog
import tkinter as tk
from tkinter import font, messagebox
from tkinter import ttk
import serial
import serial.tools.list_ports
import openpyxl
from datetime import datetime

# Color scheme
txt_color = "#a5a5a5"
bg_color = "#0a0a0a"
field_bg = "#1C1C1C"
active_col = "#336C33"

class SerialLoggerApp:
    def __init__(self, master):
        self.master = master
        master.title("STM32 Data Logger")
        master.geometry("1100x850")
        master.configure(bg=bg_color)

        # Serial Configuration Frame
        control_frame = tk.LabelFrame(master, text="Serial Configuration", bg=bg_color, fg=txt_color,
                                      font=("Arial",12,"bold"), bd=2, relief="flat")
        control_frame.pack(fill="x", padx=12, pady=6)

        # COM Port
        tk.Label(control_frame, text="COM Port:", bg=bg_color, fg=txt_color).grid(row=0, column=0,
                                                                                     padx=6, pady=4, sticky='w')
        self.port_var = tk.StringVar()
        self.port_combo = ttk.Combobox(control_frame, textvariable=self.port_var,
                                       width=12, state="readonly")
        self.port_combo.grid(row=0, column=1, padx=6)
        self.refresh_btn = tk.Button(control_frame, text="Refresh", command=self.refresh_ports,
                                     bg=field_bg, fg=txt_color, activebackground=active_col, bd=0)
        self.refresh_btn.grid(row=0, column=2, padx=6)
        self.refresh_ports()  # populate initially

        # Baud Rate
        tk.Label(control_frame, text="Baud Rate:", bg=bg_color, fg=txt_color).grid(row=0, column=3, padx=6)
        self.baud_var = tk.StringVar(value="115200")
        tk.Entry(control_frame, textvariable=self.baud_var, width=8,
                 bg=field_bg, fg="#ffffff", insertbackground="#3b3b3b").grid(row=0, column=4, padx=6)

        # Connect Button
        self.connect_btn = tk.Button(control_frame, text="Connect", command=self.connect_serial,
                                     bg=field_bg, fg=txt_color, activebackground=active_col, bd=0)
        self.connect_btn.grid(row=0, column=5, padx=8)

        # Trigger Button
        self.trigger_btn = tk.Button(control_frame, text="Trigger", command=self.send_trigger,
                                     bg=field_bg, fg=txt_color, activebackground=active_col, bd=0)
        self.trigger_btn.grid(row=0, column=6, padx=8)

        # Mode Select Dropdown
        tk.Label(control_frame, text="Mode:", bg=bg_color, fg=txt_color).grid(row=0, column=7, padx=6)
        self.mode_var = tk.StringVar(value="Continuously")
        self.mode_combo = ttk.Combobox(control_frame, textvariable=self.mode_var,
                                       values=["Continuously", "Sweep"], width=11, state="readonly")
        self.mode_combo.grid(row=0, column=8, padx=6)

         # Add Average Button (for post-processing)
        self.avg_btn = tk.Button(control_frame, text="Average Sweeps", command=self.average_sweeps,
                                 bg=field_bg, fg=txt_color, activebackground=active_col, bd=0)
        self.avg_btn.grid(row=0, column=9, padx=8)


        # Status Label
        self.output_label = tk.Label(master, text="Awaiting connection...", font=("Arial",11),
                                     fg=txt_color, bg=bg_color)
        self.output_label.pack(pady=4)

        # Recording Controls
        record_frame = tk.LabelFrame(master, text="Recording Controls", bg=bg_color, fg=txt_color,
                                     font=("Arial",12,"bold"), bd=2, relief="flat")
        record_frame.pack(fill="x", padx=12, pady=6)

        self.start_rec_btn = tk.Button(record_frame, text="Start Recording",
                                       command=self.start_recording, state="disabled",
                                       bg=field_bg, fg=txt_color, activebackground=active_col, bd=0)
        self.start_rec_btn.grid(row=0, column=0, padx=10, pady=4)

        self.stop_rec_btn = tk.Button(record_frame, text="Stop & Save", command=self.stop_recording,
                                      state="disabled", bg=field_bg, fg=txt_color,
                                      activebackground=active_col, bd=0)
        self.stop_rec_btn.grid(row=0, column=1, padx=10, pady=4)

        self.indicator = tk.Label(record_frame, text="●", font=("Arial",40,"bold"), fg=bg_color,
                                  bg=bg_color)
        self.indicator.grid(row=0, column=2, padx=10)
        self.ind_visible = False

        tk.Label(record_frame, text="Save To:", bg=bg_color, fg=txt_color).grid(row=0, column=3, padx=6)
        self.save_path_var = tk.StringVar(value=os.getcwd())
        tk.Entry(record_frame, textvariable=self.save_path_var, width=30,
                 bg=field_bg, fg="#ffffff", insertbackground="#ffffff", bd=0).grid(row=0, column=4, padx=6)
        tk.Button(record_frame, text="Browse…", command=self.browse_folder,
                  bg=field_bg, fg=txt_color, activebackground=active_col, bd=0).grid(row=0, column=5, padx=6)
        
        #sweep tracking
        self.sweep_index = 1
        self.sweep_label = tk.Label(record_frame, text=f"Sweep: {self.sweep_index}", bg=bg_color, fg=txt_color)
        self.sweep_label.grid(row=0, column=10, padx=6)

        # Reset sweep counter
        self.reset_btn = tk.Button(record_frame, text="Reset", command=self.reset_sweep,
                                   bg=field_bg, fg=txt_color, activebackground=active_col, bd=0)
        self.reset_btn.grid(row=0, column=11, padx=6)

        #sweep input
        tk.Label(record_frame, text="N:", bg=bg_color, fg=txt_color).grid(row=0, column=12, padx=6)
        self.sweep_input = tk.IntVar(value=10)
        tk.Entry(record_frame, textvariable=self.sweep_input, width=8,
                 bg=field_bg, fg="#ffffff", insertbackground="#3b3b3b").grid(row=0, column=13, padx=6)

        # PWM Progress Bar
        pwm_frame = tk.LabelFrame(master, text="Analog OUT", bg=bg_color, fg=txt_color,
                                  font=("Arial",12,"bold"), bd=2, relief="flat")
        pwm_frame.pack(fill="x", padx=12, pady=6)
        self.pwm_var = tk.IntVar()
        style = ttk.Style()
        style.theme_use("clam")
        style.configure("green.Horizontal.TProgressbar", troughcolor=bg_color, background="#e4d100")
        self.pwm_bar = ttk.Progressbar(pwm_frame, style="green.Horizontal.TProgressbar",
                                       maximum=238, variable=self.pwm_var)
        self.pwm_bar.pack(fill="x", padx=10, pady=8)

        # Channel Displays CH2..CH9 and CH10
        self.value_frame = tk.Frame(master, bg=bg_color)
        self.value_frame.pack(padx=12, pady=10)
        self.custom_vars = []
        self.value_labels = []
        self.seg_font = font.Font(family="LEMON-MILK", size=40, weight="bold")

        default_labels = ["CH1","CH2","CH3","CH4","CH5","CH6","Analog in","SPEED","COUNT"]
        default_vals   = ["0.00 V","0.00 V","0.00 V","0.00 V","0.0 V","0 V","0","0 Hz","0"]

        for i in range(1,10):
            col = 0 if i < 6 else 1
            row = (i-1) % 5
            cell = tk.Frame(self.value_frame, bg="#000000", padx=8, pady=8,
                            bd=2, relief="ridge", highlightbackground="#191919",
                            highlightthickness=2)
            cell.grid(row=row*2, column=col, padx=20, pady=5, sticky='w')
            var = tk.StringVar(value=default_labels[i-1])
            entry = tk.Entry(cell, textvariable=var, font=("Arial",20), width=12,
                             bg=bg_color, fg="#ffffff", insertbackground=bg_color, bd=0)
            entry.pack(side='left', padx=(0,8))
            entry.bind('<Return>', self.userparam_trigger)
            self.custom_vars.append(var)
            val_lbl = tk.Label(cell, text=default_vals[i-1], font=self.seg_font,
                               fg="#ff0000", bg="#000000", width=6,
                               anchor='e', justify='right')
            val_lbl.pack(side='left')
            self.value_labels.append(val_lbl)

        cell10 = tk.Frame(self.value_frame, bg="#000000", padx=8, pady=8,
                          bd=2, relief="ridge", highlightbackground="#191919",
                          highlightthickness=2)
        cell10.grid(row=8, column=1, padx=20, pady=5, sticky='w')
        self.user10_var = tk.StringVar(value="UserParam")
        lbl10 = tk.Entry(cell10, textvariable=self.user10_var, font=("Arial",20), width=12,
                          bg=bg_color, fg="#ffffff", insertbackground=bg_color, bd=0)
        lbl10.pack(side='left', padx=(0,10))
        lbl10.bind('<Return>', self.userparam_trigger)
        self.custom_vars.append(self.user10_var)
        self.user10_value_var = tk.IntVar(value=0)
        val10 = tk.Entry(cell10, textvariable=self.user10_value_var, font=self.seg_font,
                          fg="#ff0000", bg="#000000", insertbackground="#ff0000", width=7, bd=0)
        val10.pack(side='left')
        val10.bind('<Return>', self.userparam_trigger)
        self.lbl10 = lbl10
        self.val10 = val10
        self.cursor_visible = True
        self.blink_cursor()

        # State flags
        self.connected = False
        self.recording = False
        self.ser = None

        # Sweep tracking for Sweep mode
        self.sweep_index = 1
        self.last_pwm = -1
        self.current_sheet = None
        self.default_headers = None
        self.user_headers = None

    def reset_sweep(self):
        """Reset the sweep counter back to zero and update its label"""
        self.sweep_index = 0
        self.sweep_label.config(text=f"Sweep: {self.sweep_index}")

    def refresh_ports(self):
        ports = [p.device for p in serial.tools.list_ports.comports()]
        self.port_combo['values'] = ports
        if ports:
            self.port_var.set(ports[0])

    def browse_folder(self):
        folder = filedialog.askdirectory(initialdir=self.save_path_var.get(), title="Select Save Folder")
        if folder:
            self.save_path_var.set(folder)

    def send_trigger(self):
        if self.connected:
            try:
                self.ser.write(b't')
            except:
                pass

    def userparam_trigger(self, event=None):
        if self.connected:
            try:
                self.ser.write(b't')
            except:
                pass

    def blink_indicator(self):
        if not self.recording:
            self.indicator.config(fg=bg_color)
            return
        self.ind_visible = not self.ind_visible
        color = "#ff0000" if self.ind_visible else bg_color
        self.indicator.config(fg=color)
        self.master.after(500, self.blink_indicator)

    def blink_cursor(self):
        self.cursor_visible = not self.cursor_visible
        color = "#ffffff" if self.cursor_visible else bg_color
        self.lbl10.config(insertbackground=color)
        self.val10.config(insertbackground=color)
        self.master.after(500, self.blink_cursor)

    def connect_serial(self):
        try:
            port, baud = self.port_var.get(), int(self.baud_var.get())
            self.ser = serial.Serial(port, baud, timeout=0.1)
            self.connected = True
            self.connect_btn.config(state="disabled")
            self.start_rec_btn.config(state="normal")
            self.output_label.config(text=f"Connected {port}@{baud}")
            # start non-blocking polling
            self.master.after(0, self.read_serial)
        except Exception as e:
            messagebox.showerror("Connection Error", str(e))

    def start_recording(self):
        if not self.connected:
            return
        self.recording = True
        self.start_rec_btn.config(state="disabled")
        self.stop_rec_btn.config(state="normal")

        self.default_headers = ["Time", "Time_MCU"] + [f"CH{i}" for i in range(1,10)] + ["_", "PWM"]
        self.user_headers = ["Time", "Time_MCU"] + [var.get() for var in self.custom_vars] + ["PWM"]

        self.wb = openpyxl.Workbook()
        if self.mode_var.get() == "Sweep":
            default = self.wb.active
            self.wb.remove(default)
            #self.sweep_index = 1
            self.current_sheet = self.wb.create_sheet(title=f"Sweep_{self.sweep_index}")
            self.current_sheet.append(self.default_headers)
            self.current_sheet.append(self.user_headers)
            self.last_pwm = -1
        else:
            ws = self.wb.active
            ws.title = "Data"
            ws.append(self.default_headers)
            ws.append(self.user_headers)

        self.blink_indicator()
        self.output_label.config(text="Recording...")

    def stop_recording(self):
        if not self.recording:
            return
        self.recording = False
        self.start_rec_btn.config(state="normal")
        self.stop_rec_btn.config(state="disabled")
        self.connect_btn.config(state="normal")

        folder = self.save_path_var.get()
        fname = f"STM32_Log_{datetime.now():%Y%m%d_%H%M%S}.xlsx"
        fullpath = os.path.join(folder, fname)
        self.wb.save(fullpath)
        self.output_label.config(text=f"Saved: {fullpath}")
        self.indicator.config(fg=bg_color)
        self.reset_sweep()

    def read_serial(self):
        if not self.connected:
            return

        try:
            line = self.ser.readline().decode(errors='ignore').strip()
            if line:
                parts = [p for p in line.split(',') if p]
                if len(parts) >= 11:
                    raw = list(map(int, parts))

                    # Update display
                    disp_vals = (
                        [f"{raw[i]/1000:.2f} V" for i in range(1,5)] +
                        [f"{raw[5]/1000:.1f} V", f"{raw[6]//1000} V",
                         str(raw[7]), f"{raw[8]} Hz", str(raw[9])]
                    )
                    for lbl, val in zip(self.value_labels, disp_vals):
                        lbl.config(text=val)
                    pwm = raw[10]
                    self.pwm_var.set(pwm)
                    self.output_label.config(text=f"Last update: {datetime.now():%H:%M:%S}")

                    # Record to Excel
                    if self.recording:
                        ts = datetime.now().strftime("%H:%M:%S.%f")[:-3]
                        numeric_vals = [
                            raw[0],
                            raw[1]/1000.0, raw[2]/1000.0,
                            raw[3]/1000.0, raw[4]/1000.0,
                            raw[5]/1000.0, raw[6]//1000,
                            raw[7], raw[8], raw[9],
                        ] + [self.user10_value_var.get(), pwm]

                        if self.mode_var.get() == "Sweep":
                            self.current_sheet.append([ts] + numeric_vals)
                            if pwm==238:
                                
                                self.sweep_index += 1
                                self.sweep_label.config(text=f"Sweep: {self.sweep_index}")
                                self.current_sheet = self.wb.create_sheet(title=f"Sweep_{self.sweep_index}")
                                self.current_sheet.append(self.default_headers)
                                self.current_sheet.append(self.user_headers)
                               
                                
                                self.last_pwm = pwm
                                if self.sweep_index < self.sweep_input.get():
                                    self.send_trigger()
                            
                            
                        else:
                            self.wb.active.append([ts] + numeric_vals)
        except Exception as e:
            print("Read Error:", e)
            self.connected = False
            self.connect_btn.config(state="normal")
            self.output_label.config(text="Disconnected")

        # schedule next read
        self.master.after(10, self.read_serial)
    def average_sweeps(self):
        try:
            file_path = filedialog.askopenfilename(initialdir=self.save_path_var.get(),
                                                   title="Select Excel File",
                                                   filetypes=[("Excel Files", "*.xlsx")])
            if not file_path:
                return

            wb = openpyxl.load_workbook(file_path)
            sweep_sheets = [ws for ws in wb.worksheets if ws.title.startswith("Sweep_")]
            if not sweep_sheets:
                messagebox.showinfo("No Sweeps", "No sheets named 'Sweep_' found.")
                return

            all_data = []
            for ws in sweep_sheets:
                rows = []
                for row in ws.iter_rows(min_row=3, values_only=True):  # skip header rows
                    # skip timestamp (col 1), only numeric columns
                    vals = list(row)[1:]
                    rows.append(vals)
                all_data.append(rows)

            lengths = {len(sheet) for sheet in all_data}
            if len(lengths) != 1:
                messagebox.showerror("Mismatch", "Sweep sheets do not have equal row counts.")
                return

            averaged_data = []
            for i in range(lengths.pop()):
                cols = zip(*(sheet[i] for sheet in all_data))
                averaged_row = [sum(col)/len(col) for col in cols]
                averaged_data.append(averaged_row)

            if "Averaged" in wb.sheetnames:
                del wb["Averaged"]
            ws_out = wb.create_sheet("Averaged")

            for row in sweep_sheets[0].iter_rows(min_row=1, max_row=2, values_only=True):
                ws_out.append(list(row))

                        # write averaged rows
            # Each row in averaged_data excludes the timestamp string (we skip col0)
            # So we reconstruct rows as: blank or original timestamp for col0, then averaged numeric for rest
            for idx, row_vals in enumerate(averaged_data, start=3):
                # get original timestamp from first sheet, row idx
                ts = sweep_sheets[0].cell(row=idx, column=1).value
                # averaged_data entry has length = num_columns -1
                ws_out.append([ts] + row_vals)

            wb.save(file_path)
            self.output_label.config(text=f"Averaged data saved in 'Averaged' sheet")
            self.output_label.config(text=f"Averaged data saved in 'Averaged' sheet")
        except Exception as e:
            messagebox.showerror("Averaging Error", str(e))


if __name__ == '__main__':
    root = tk.Tk()
    app = SerialLoggerApp(root)
    root.mainloop()
